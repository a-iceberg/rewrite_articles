import os
import json
import re
import pandas as pd
from openai import OpenAI
from bs4 import BeautifulSoup
import tiktoken
from tqdm import tqdm
from openpyxl import load_workbook
import time


auth_path = "data/auth.json"
with open(auth_path, "r") as f:
    config = json.load(f)

OPENAI_API_KEY = config["OPENAI_API_KEY"]


client = OpenAI(api_key=OPENAI_API_KEY)


prompt_path = "data/prompt.txt"
with open(prompt_path, "r", encoding="utf-8") as f:
    base_prompt = f.read().strip()


input_file_name = "data/NEWS ICEBERG.xlsx"
output_file_name = "data/Processed_articles.xlsx"  

data = pd.read_excel(input_file_name)


if not os.path.exists(output_file_name):
    df_empty = pd.DataFrame(columns=["id", "title", "url", "processed_text", "description"])
    df_empty.to_excel(output_file_name, index=False)
    print(f"[INFO] Файл '{output_file_name}' не существует. Создан пустой файл.")


model_name = "gpt-4o-2024-11-20"
INPUT_COST_PER_M = 2.50
OUTPUT_COST_PER_M = 10.00

# Функция подсчёта токенов
def count_tokens(text, model=model_name):

    #text = text.strip()
    encoder = tiktoken.encoding_for_model(model)
    return len(encoder.encode(text))  


# Функция расчёта стоимости токенов
def calculate_cost(tokens, input=True):
    tokens_in_millions = tokens / 1_000_000
    return tokens_in_millions * (INPUT_COST_PER_M if input else OUTPUT_COST_PER_M)

# Функция очистки HTML
def clean_html(text):
    soup = BeautifulSoup(text, "html.parser")
    

    for style in soup.find_all("style"):
        style.decompose()


    article_content = f"<article>{soup.get_text()}</article>"
    return article_content.strip()

# Функция очистки HTML-ответов от Markdown-обертки
def clean_html_response(response: str):
    return re.sub(r"```html\s*([\s\S]*?)\s*```", r"\1", response).strip()

# Функция извлечения содержимого article
def extract_article_content(html_text):
    soup = BeautifulSoup(html_text, "html.parser")
    article_tag = soup.find("article")
    if article_tag:
        return article_tag.decode_contents().strip()
    return "Ошибка: article не найден"

# Функция сохранения данных в Excel
def save_to_excel(data, file_name):
    file_exists = os.path.exists(file_name)
    
    df = pd.DataFrame(data, columns=["id", "title", "url", "processed_text", "description"])
    
    try:
        with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            if file_exists:
                workbook = load_workbook(file_name)
                sheet = workbook.active
                start_row = sheet.max_row
            else:
                start_row = 0

            df.to_excel(writer, index=False, header=start_row == 0, startrow=start_row)
        
        print(f"[INFO] Сохранено {len(data)} статей в {file_name}")
    except Exception as e:
        print(f"[ERROR] Ошибка сохранения в Excel: {e}")

# Функция обработки статьи GPT-4o с прогресс-баром
def process_article(data):
    print(f"[INFO] Обрабатываем статьи...")

    results = []
    total_tokens = 0
    total_cost = 0


    with tqdm(total=len(data), desc="Обработка статей", unit="статья") as pbar:
        for idx, row in data.iterrows():  
            article_id, url, text = row["id"], row["url"], row["text"]
            formatted_text = clean_html(text)
            input_tokens = count_tokens(formatted_text)

            try:
                response = client.chat.completions.create(
                    model=model_name,
                    temperature=0.3,
                    max_tokens=2000,
                    messages=[
                        {"role": "system", "content": base_prompt},
                        {"role": "user", "content": formatted_text}
                    ]
                )
                
                model_response = response.choices[0].message.content.strip()
                model_response = clean_html_response(model_response) 
                output_tokens = count_tokens(model_response)
                #print(input_tokens, output_tokens)


                total_tokens += output_tokens

                total_cost += calculate_cost(input_tokens, input=True) + calculate_cost(output_tokens, input=False)
                #print(total_cost)
                soup = BeautifulSoup(model_response, "html.parser")
                
                # Извлекаем title
                meta_title = soup.find("title").text if soup.find("title") else "Без заголовка"

                # Извлекаем description
                meta_description = soup.find("meta", {"name": "description"})
                meta_description = meta_description["content"] if meta_description else ""

                # Извлекаем только <article>...</article>
                processed_html = extract_article_content(model_response)

                results.append([article_id, meta_title, url, processed_html, meta_description])

            except Exception as e:
                print(f"[ERROR] Ошибка при обработке статьи {article_id}: {e}")
                results.append([article_id, "Ошибка", "Ошибка", "<article>Ошибка обработки</article>", "Ошибка"])


            pbar.update(1)

            # Сохраняем каждые 20 статей
            if (idx + 1) % 20 == 0:
                save_to_excel(results, output_file_name)
                results = []  

    return results, total_tokens, total_cost


data = pd.read_excel(input_file_name, sheet_name=0)


processed_results, total_tokens, total_cost = process_article(data) 


if processed_results:
    save_to_excel(processed_results, output_file_name)

print(f"[INFO] Обработанные статьи сохранены в {output_file_name}")
print(f"[INFO] Итоговая стоимость: ${total_cost:.4f}")
print(f"[INFO] Итоговое количество токенов: {total_tokens}")
